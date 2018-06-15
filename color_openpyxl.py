'''color_openpyxl颜色序号生成列表'''
from openpyxl.workbook import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill

#新建一个excel文件
wb = Workbook()
#使用激活中的sheet，就是sheet1
ws = wb.active


#颜色参数元组
color = colors.COLOR_INDEX
for i in range(64):
    v = color[i]
    ws['A%d'%(i+1)] = v
    ws['A%d'%(i+1)].fill = PatternFill(fill_type='solid', fgColor=v)


wb.save('color_openpyxl.xlsx')