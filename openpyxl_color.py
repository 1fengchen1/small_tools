from openpyxl.workbook import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active

ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = True



color = colors.COLOR_INDEX
print(color)
for i in range(64):
    v = color[i]
    print(v)
    ws['A%d'%(i+1)] = v
    ws['A%d'%(i+1)].fill = PatternFill(fill_type='solid', fgColor=v)


wb.save('color_openpyxl1.xlsx')